# ============================================================
#   SISTEMA DE REGISTRO DE ASISTENCIA SEMANAL
# ============================================================

from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NOMBRES_DIAS = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sabado"]

# ============================================================
# PASO 1: Registrar datos del usuario
# Responsable: Sesarina
# ============================================================

print("=" * 50)
print("   REGISTRO DE ASISTENCIA SEMANAL")
print("=" * 50)
print()

codigo    = input("Codigo de usuario : ")
nombre    = input("Nombre completo   : ")
area      = input("Area              : ")
turno     = input("Turno             : ")
aprobador = input("Aprobador         : ")

print()

# ============================================================
# PASO 2: Registrar fecha de inicio de semana
# Responsable: Brian
# ============================================================

fecha_lunes = input("Fecha del LUNES de la semana (DD/MM/AAAA): ")
lunes = datetime.strptime(fecha_lunes, "%d/%m/%Y")

fechas = []
for i in range(6):
    fechas.append(lunes + timedelta(days=i))

print()

# ============================================================
# PASO 3: Validar dias trabajados
# Responsable: Omar
# ============================================================

print("-" * 50)
print("  REGISTRO DIARIO")
print("-" * 50)
print()

lista_dia     = []
lista_fecha   = []
lista_trabajo = []
lista_inicio  = []
lista_fin     = []
lista_horas   = []

for i in range(6):
    nombre_dia = NOMBRES_DIAS[i]
    fecha_dia  = fechas[i].strftime("%d/%m/%Y")

    print("Dia: " + nombre_dia + " " + fecha_dia)
    respuesta = input("  Trabajaste? (s/n): ")

# ============================================================
# PASO 4: Calcular horas trabajadas por dia
# Responsable: Sesarina
# ============================================================

    if respuesta == "s":
        inicio = input("  Hora de inicio (HH:MM): ")
        fin    = input("  Hora de fin    (HH:MM): ")

        hora_i = datetime.strptime(inicio, "%H:%M")
        hora_f = datetime.strptime(fin,    "%H:%M")
        diferencia = hora_f - hora_i

    if diferencia.total_seconds() < 0:
       diferencia = diferencia + timedelta(days=1)

        horas_trabajadas = diferencia.total_seconds() / 3600

        lista_dia.append(nombre_dia)
        lista_fecha.append(fecha_dia)
        lista_trabajo.append("Si")
        lista_inicio.append(inicio)
        lista_fin.append(fin)
        lista_horas.append(round(horas_trabajadas, 2))
    else:
        lista_dia.append(nombre_dia)
        lista_fecha.append(fecha_dia)
        lista_trabajo.append("No")
        lista_inicio.append("--")
        lista_fin.append("--")
        lista_horas.append(0)

    print()

# ============================================================
# PASO 5: Calcular total horas semanales
# Responsable: Brian
# ============================================================

total_horas = 0
for i in range(6):
    total_horas = total_horas + lista_horas[i]


# ============================================================
# PASO 6: Calcular horas extras
# Responsable: Omar
# ============================================================

horas_extras = total_horas - 48
if horas_extras < 0:
    horas_extras = 0

# ============================================================
# PASO 7: Contar dias trabajados
# Responsable: Sesarina
# ============================================================

dias_trabajados = 0
for i in range(6):
    if lista_trabajo[i] == "Si":
        dias_trabajados = dias_trabajados + 1
        
# ============================================================
# PASO 8: Determinar tipo de jornada
# Responsable: Brian
# ============================================================

if dias_trabajados >= 6:
    tipo_jornada = "Completa"
else:
    tipo_jornada = "Incompleta"


# ============================================================
# PASO 9: Generar tabla de asistencia (consolidado en pantalla)
# Responsable: Omar
# ============================================================

def fmt_horas(h):
    hh = int(h)
    mm = int(round((h - hh) * 60))
    return str(hh) + "h " + str(mm).zfill(2) + "m"

print()
print("=" * 60)
print("            CONSOLIDADO SEMANAL")
print("=" * 60)
print("Codigo    :", codigo)
print("Nombre    :", nombre)
print("Area      :", area)
print("Turno     :", turno)
print("Aprobador :", aprobador)
print("-" * 60)
print("{:<12} {:<13} {:<8} {:<8} {:<8}".format("Dia", "Fecha", "Inicio", "Fin", "Horas"))
print("-" * 60)

for i in range(6):
    if lista_trabajo[i] == "Si":
        print("{:<12} {:<13} {:<8} {:<8} {:<8}".format(
            lista_dia[i], lista_fecha[i],
            lista_inicio[i], lista_fin[i],
            fmt_horas(lista_horas[i])))
    else:
        print("{:<12} {:<13} {}".format(
            lista_dia[i], lista_fecha[i], "No trabajo"))

print("=" * 60)
print("Dias trabajados  :", dias_trabajados, "de 6")
print("Total horas      :", fmt_horas(total_horas))
print("Horas extras     :", fmt_horas(horas_extras))
print("Tipo de jornada  :", tipo_jornada)
print("=" * 60)

# ============================================================
# PASO 10: Exportar reporte Excel
# Responsable: Omar, Brian y Sesarina
# ============================================================

# --- Colores del diseño ---
AZUL_OSC  = "1F4E79"
AZUL_CLAR = "D6E4F0"
GRIS      = "2C3E50"
VERDE     = "1E8449"
ROJO      = "C0392B"
AMARILLO  = "FFF2CC"
BLANCO    = "FFFFFF"
NEGRO     = "000000"
BLANCO_F  = "FFFFFF"

def estilo(celda, negrita=False, fondo=None, color_fuente=NEGRO,
           alinear="left", size=11):
    celda.font      = Font(bold=negrita, color=color_fuente,
                           size=size, name="Arial")
    celda.alignment = Alignment(horizontal=alinear, vertical="center",
                                wrap_text=True)
    if fondo:
        celda.fill = PatternFill("solid", start_color=fondo)
    lado = Side(style="thin")
    celda.border = Border(left=lado, right=lado, top=lado, bottom=lado)

def combinar(ws, f1, c1, f2, c2, valor, negrita=False, fondo=None,
             color_fuente=BLANCO_F, alinear="center", size=11):
    ws.merge_cells(start_row=f1, start_column=c1,
                   end_row=f2,   end_column=c2)
    c = ws.cell(row=f1, column=c1, value=valor)
    estilo(c, negrita=negrita, fondo=fondo,
           color_fuente=color_fuente, alinear=alinear, size=size)

libro = Workbook()
hoja  = libro.active
hoja.title = "Asistencia Semanal"

hoja.column_dimensions["A"].width = 3
hoja.column_dimensions["B"].width = 14
hoja.column_dimensions["C"].width = 13
hoja.column_dimensions["D"].width = 12
hoja.column_dimensions["E"].width = 9
hoja.column_dimensions["F"].width = 9
hoja.column_dimensions["G"].width = 10
hoja.column_dimensions["H"].width = 12
hoja.column_dimensions["I"].width = 12.3

hoja.row_dimensions[1].height = 40
combinar(hoja, 1,2, 1,8, "TABLA DE ASISTENCIA SEMANAL",
         negrita=True, fondo=AZUL_OSC, size=16)

hoja.row_dimensions[2].height = 18
fecha_gen = datetime.now().strftime("%d/%m/%Y %H:%M")
combinar(hoja, 2,2, 2,8, "Generado: " + fecha_gen,
         fondo=GRIS, size=9, alinear="right")

combinar(hoja, 4,2, 4,8, "DATOS DEL COLABORADOR",
         negrita=True, fondo=GRIS, size=11)

datos_usuario = [
    ("Codigo",    codigo),
    ("Nombre",    nombre),
    ("Area",      area),
    ("Turno",     turno),
    ("Aprobador", aprobador),
]
for idx, (etiqueta, valor) in enumerate(datos_usuario):
    fila = 5 + idx
    hoja.row_dimensions[fila].height = 20
    c_etiq = hoja.cell(row=fila, column=2, value=etiqueta)
    estilo(c_etiq, negrita=True, fondo=AZUL_CLAR,
           color_fuente=NEGRO, alinear="right")
    combinar(hoja, fila,3, fila,8, valor,
             fondo=BLANCO, color_fuente=NEGRO, alinear="left")

combinar(hoja, 11,2, 11,8, "REGISTRO DIARIO",
         negrita=True, fondo=GRIS, size=11)

hoja.row_dimensions[12].height = 25
cabeceras = ["N°", "Dia", "Fecha", "Trabajo?",
             "H. Inicio", "H. Fin", "Horas", "Observacion"]
for col, cab in enumerate(cabeceras, 2):
    c = hoja.cell(row=12, column=col, value=cab)
    estilo(c, negrita=True, fondo=AZUL_OSC,
           color_fuente=BLANCO_F, alinear="center", size=10)

for i in range(6):
    fila = 13 + i
    hoja.row_dimensions[fila].height = 20

    if i % 2 == 0:
        fondo_fila = AZUL_CLAR
    else:
        fondo_fila = BLANCO

    if lista_trabajo[i] == "Si":
        color_trabajo = VERDE
        texto_trabajo = "Si"
        observacion   = ""
    else:
        color_trabajo = ROJO
        texto_trabajo = "No"
        observacion   = "Ausente"

    hoja_horas = fmt_horas(lista_horas[i])

    datos_fila = [
        (2, i + 1,              NEGRO,         "center", False),
        (3, lista_dia[i],       NEGRO,         "left",   False),
        (4, lista_fecha[i],     NEGRO,         "center", False),
        (5, texto_trabajo,      color_trabajo, "center", True ),
        (6, lista_inicio[i],    NEGRO,         "center", False),
        (7, lista_fin[i],       NEGRO,         "center", False),
        (8, hoja_horas,         NEGRO,         "center", False),
        (9, observacion,        NEGRO,         "center", False),
    ]

    for col, valor, color_f, alin, negr in datos_fila:
        c = hoja.cell(row=fila, column=col, value=valor)
        estilo(c, negrita=negr, fondo=fondo_fila,
               color_fuente=color_f, alinear=alin)

combinar(hoja, 21,2, 21,8, "RESUMEN DE HORAS",
         negrita=True, fondo=GRIS, size=11)

hoja.row_dimensions[22].height = 22
hoja.row_dimensions[23].height = 22
hoja.row_dimensions[24].height = 22

resumen_datos = [
    (22, "Dias trabajados",        str(dias_trabajados) + " de 6", AZUL_CLAR),
    (23, "Total horas semanales",  fmt_horas(total_horas),         AZUL_CLAR),
    (24, "Horas extras (48h)",     fmt_horas(horas_extras),        AMARILLO),
]

for fila, etiqueta, valor, fondo_res in resumen_datos:
    combinar(hoja, fila,2, fila,5, etiqueta,
             negrita=True, fondo=fondo_res,
             color_fuente=NEGRO, alinear="left")
    combinar(hoja, fila,6, fila,8, valor,
             negrita=True, fondo=fondo_res,
             color_fuente=NEGRO, alinear="center")

combinar(hoja, 26,2, 26,8, "TIPO DE JORNADA",
         negrita=True, fondo=GRIS, size=11)

hoja.row_dimensions[27].height = 30
if tipo_jornada == "Completa":
    color_jornada = VERDE
    desc_jornada  = "El colaborador cumplio 6 o mas dias trabajados en la semana."
else:
    color_jornada = ROJO
    desc_jornada  = "El colaborador trabajo menos de 6 dias en la semana."

combinar(hoja, 27,2, 27,8, "JORNADA " + tipo_jornada.upper(),
         negrita=True, fondo=color_jornada, size=14)

hoja.row_dimensions[28].height = 18
combinar(hoja, 28,2, 28,8, desc_jornada,
         fondo=AZUL_CLAR, color_fuente=NEGRO,
         alinear="center", size=9)

combinar(hoja, 30,2, 30,4, "Firma del colaborador",
         fondo=BLANCO, color_fuente=NEGRO, alinear="center", size=9)
combinar(hoja, 30,6, 30,8, "Firma del aprobador",
         fondo=BLANCO, color_fuente=NEGRO, alinear="center", size=9)

combinar(hoja, 33,2, 33,4, "_________",
         fondo=BLANCO, color_fuente=NEGRO, alinear="center")
combinar(hoja, 33,6, 33,8, "_________",
         fondo=BLANCO, color_fuente=NEGRO, alinear="center")

nombre_archivo = "asistencia_" + codigo + "_" + lista_fecha[0].replace("/", "-") + ".xlsx"
libro.save(nombre_archivo)

print()
print("=" * 60)
print("  Archivo Excel generado exitosamente!")
print("  Nombre:", nombre_archivo)
print("  Se guardo en la misma carpeta donde esta este script.")
print("=" * 60)


